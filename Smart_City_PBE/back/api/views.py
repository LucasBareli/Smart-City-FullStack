from .models import *
from .serializers import *
from rest_framework.generics import ListCreateAPIView, RetrieveUpdateDestroyAPIView, CreateAPIView, ListAPIView 
from rest_framework.permissions import IsAuthenticated
from django.contrib.auth.models import User
from rest_framework.filters import SearchFilter
from django_filters.rest_framework import DjangoFilterBackend, FilterSet, DateTimeFilter
from django.utils import timezone
from datetime import timedelta
import pandas as pd
from django.views import View
from django.http import HttpResponse
from datetime import datetime
import csv
import openpyxl
from django.utils.dateparse import parse_datetime


class SignUpView(CreateAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializer

class SensoresView(ListCreateAPIView):
    queryset = Sensores.objects.all()
    serializer_class = SensoresSerializer
    permission_classes = [IsAuthenticated]

class SensoresAPIView(RetrieveUpdateDestroyAPIView):
    queryset = Sensores.objects.all()
    serializer_class = SensoresSerializer
    permission_classes = [IsAuthenticated]

class AmbientesView(ListCreateAPIView):
    queryset = Ambientes.objects.all()
    serializer_class = AmbientesSerializer
    permission_classes = [IsAuthenticated]

class AmbientesAPIView(RetrieveUpdateDestroyAPIView):
    queryset = Ambientes.objects.all()
    serializer_class = AmbientesSerializer
    permission_classes = [IsAuthenticated]

class HistoricosView(ListCreateAPIView):
    queryset = Historicos.objects.all()
    serializer_class = HistoricosSerializer
    permission_classes = [IsAuthenticated]

class HistoricosAPIView(RetrieveUpdateDestroyAPIView):
    queryset = Historicos.objects.all()
    serializer_class = HistoricosSerializer
    permission_classes = [IsAuthenticated]


# Filtragem por data inicio e data final


class HistoricosFilter(FilterSet):
    data_inicial = DateTimeFilter(field_name="timestamp", lookup_expr='gte')
    data_final = DateTimeFilter(field_name="timestamp", lookup_expr='lte')

    class Meta:
        model = Historicos
        fields = ['data_inicial', 'data_final', 'sensor', 'ambiente']


# Filtros gerais


class SensoresSearchView(ListAPIView):
    queryset = Sensores.objects.all()
    serializer_class = SensoresSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = (DjangoFilterBackend, SearchFilter)
    filterset_fields = ['status']
    search_fields = ['sensor', 'mac_address', 'unidade_med', 'latitude', 'longitude']

class AmbientesSearchView(ListAPIView):
    queryset = Ambientes.objects.all()
    serializer_class = AmbientesSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = (DjangoFilterBackend, SearchFilter)
    search_fields = ['sig', 'descricao', 'ni', 'responsavel']

class HistoricosSearchView(ListAPIView):
    queryset = Historicos.objects.all()
    serializer_class = HistoricosSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = (DjangoFilterBackend, SearchFilter)
    filterset_class = HistoricosFilter
    search_fields = [
        'sensores__sensor',
        'sensores__mac_address',
        'sensores__id',
        'ambientes__descricao',
        'ambientes__id',
        'timestamp',
        'valor'
    ]


# Filtragem Dupla

class HistoricosSensorDataView(ListAPIView):
    queryset = Historicos.objects.all()
    serializer_class = HistoricosSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = (DjangoFilterBackend,)
    
    def get_queryset(self):
        sensor = self.request.query_params.get('sensor')
        data_inicial = self.request.query_params.get('data_inicial')
        data_final = self.request.query_params.get('data_final')
        
        queryset = super().get_queryset()
        if sensor:
            queryset = queryset.filter(sensor__id=sensor)
        if data_inicial:
            queryset = queryset.filter(timestamp__gte=data_inicial)
        if data_final:
            queryset = queryset.filter(timestamp__lte=data_final)
        
        return queryset


# Filtragem tripla 

class HistoricosSensorDataHoraView(ListAPIView):
    queryset = Historicos.objects.all()
    serializer_class = HistoricosSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = (DjangoFilterBackend,)

    def get_queryset(self):
        sensor = self.request.query_params.get('sensor')
        data_inicial = self.request.query_params.get('data_inicial')
        data_final = self.request.query_params.get('data_final')

        queryset = super().get_queryset()

        if sensor:
            queryset = queryset.filter(sensor__id=sensor)
        
        if data_inicial:
            dt_inicial = parse_datetime(data_inicial)
            if dt_inicial:
                queryset = queryset.filter(timestamp__gte=dt_inicial)
        
        if data_final:
            dt_final = parse_datetime(data_final)
            if dt_final:
                queryset = queryset.filter(timestamp__lte=dt_final)

        return queryset

# Historico das ultimas 24 horas


class HistoricosUltimas24hView(ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = HistoricosSerializer

    def get_queryset(self):
        agora = timezone.now()
        limite = agora - timedelta(hours=24)
        return Historicos.objects.filter(timestamp__gte=limite).order_by('-timestamp')



# Leitura dos arquivos excel


from api.models import Sensores, Ambientes, Historicos
import os
import pandas as pd
from django.http import JsonResponse

def importar_planilhas(request):
    arquivos_planilha = {
        'contador.xlsx': ('Contador de Pessoas', 'Un'),
        'luminosidade.xlsx': ('Luminosidade', 'Lux'),
        'temperatura.xlsx': ('Temperatura', '°C'),
        'umidade.xlsx': ('Umidade', '%'),
    }

    base_path = os.path.dirname(os.path.dirname(__file__))
    sensores_inseridos = 0

    # Importar Sensores

    for arquivo, (tipo_sensor, unidade_medida) in arquivos_planilha.items():
        caminho = os.path.join(base_path, arquivo)
        try:
            dados = pd.read_excel(caminho)
        except Exception as exc:
            print(f"Erro ao carregar {arquivo}: {exc}")
            continue

        for idx, registro in dados.iterrows():
            try:
                Sensores.objects.create(
                    sensor=tipo_sensor,
                    mac_address=str(registro['mac_address']),
                    unidade_med=unidade_medida,
                    latitude=float(registro['latitude']),
                    longitude=float(registro['longitude']),
                    status=registro['status'].strip().lower() == 'ativo'
                )
                sensores_inseridos += 1
            except Exception as erro:
                print(f"Erro na linha {idx + 2} em {arquivo}: {erro}")

    print(f"Sensores inseridos: {sensores_inseridos}")

    # Importar Ambientes

    ambientes_path = os.path.join(base_path, 'Ambientes.xlsx')
    print(f"Caminho Ambientes.xlsx: {os.path.abspath(ambientes_path)}")

    if not os.path.isfile(ambientes_path):
        msg = f"Arquivo Ambientes.xlsx não encontrado em {ambientes_path}"
        print(msg)
        return JsonResponse({"erro": msg})

    try:
        ambientes_df = pd.read_excel(ambientes_path)
    except Exception as exc:
        msg = f"Erro ao abrir Ambientes.xlsx: {exc}"
        print(msg)
        return JsonResponse({"erro": msg})

    ambientes_inseridos = 0
    for idx, registro in ambientes_df.iterrows():
        try:
            Ambientes.objects.create(
                sig=str(registro['sig']),
                descricao=str(registro['descricao']),
                ni=str(registro['ni']),
                responsavel=str(registro['responsavel'])
            )
            ambientes_inseridos += 1
        except Exception as erro:
            print(f"Erro linha {idx + 1} do Ambientes.xlsx: {erro}")

    print(f"Ambientes inseridos: {ambientes_inseridos}")

    # Importar Historicos

    historicos_path = os.path.join(base_path, 'histórico.xlsx')
    print(f"Caminho histórico.xlsx: {os.path.abspath(historicos_path)}")

    if not os.path.isfile(historicos_path):
        msg = f"Arquivo histórico.xlsx não encontrado em {historicos_path}"
        print(msg)
        return JsonResponse({"erro": msg})

    try:
        historicos_df = pd.read_excel(historicos_path)
    except Exception as exc:
        msg = f"Erro ao abrir histórico.xlsx: {exc}"
        print(msg)
        return JsonResponse({"erro": msg})

    historicos_inseridos = 0
    for idx, registro in historicos_df.iterrows():
        try:
            sensor_obj = Sensores.objects.get(id=registro['sensor'])
            ambiente_obj = Ambientes.objects.get(id=registro['ambiente'])  
            
            Historicos.objects.create(
                sensor=sensor_obj,
                ambiente=ambiente_obj,
                valor=float(registro['valor']),
                timestamp=str(registro['timestamp'])
            )
            historicos_inseridos += 1
        except Sensores.DoesNotExist:
            print(f"Erro: Sensor com ID {registro['sensor']} não encontrado na linha {idx + 1}.")
        except Ambientes.DoesNotExist:
            print(f"Erro: Ambiente com ID {registro['ambiente']} não encontrado na linha {idx + 1}.")
        except Exception as erro:
            print(f"Erro linha {idx + 1} do histórico.xlsx: {erro}")

        print(f"Históricos inseridos: {historicos_inseridos}")

    return JsonResponse({
        "sucesso": f"{sensores_inseridos} sensores, {ambientes_inseridos} ambientes e {historicos_inseridos} históricos inseridos."
    })


# Exportação de arquivo xlsx

class ExportHistoricosView(View):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        # Filtrar os dados com base nos parâmetros da URL
        queryset = HistoricosFilter(request.GET, queryset=Historicos.objects.all()).qs

        # Formato do arquivo (csv ou excel)
        file_format = request.GET.get('format', 'csv').lower()

        if file_format == 'csv':
            return self.export_csv(queryset)
        elif file_format == 'excel':
            return self.export_excel(queryset)
        else:
            return HttpResponse("Formato inválido. Use 'csv' ou 'excel'.", status=400)

    def export_csv(self, queryset):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="historicos_{datetime.now().strftime("%Y-%m-%d")}.csv"'

        writer = csv.writer(response)
        # Cabeçalhos do CSV
        writer.writerow(['ID', 'Sensor', 'Ambiente', 'Valor', 'Timestamp'])

        # Dados
        for historico in queryset:
            writer.writerow([
                historico.id,
                historico.sensor.sensor,  
                historico.ambiente.descricao,  
                historico.valor,
                historico.timestamp
            ])

        return response

    def export_excel(self, queryset):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Históricos"

        # Cabeçalhos
        headers = ['ID', 'Sensor', 'Ambiente', 'Valor', 'Timestamp']
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

        # Dados
        for row_num, historico in enumerate(queryset, start=2):
            sheet.cell(row=row_num, column=1, value=historico.id)
            sheet.cell(row=row_num, column=2, value=historico.sensor.sensor)
            sheet.cell(row=row_num, column=3, value=historico.ambiente.descricao)
            sheet.cell(row=row_num, column=4, value=historico.valor)
            sheet.cell(row=row_num, column=5, value=historico.timestamp)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="historicos_{datetime.now().strftime("%Y-%m-%d")}.xlsx"'

        workbook.save(response)
        return response